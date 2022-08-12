import timeit

import pandas as pd
from openpyxl import load_workbook

from src.rules.interval_knapsack import interval_knapsack_projects
from src.rules.reverse_interval_knapsack import interval_knapsack_projects_reversed
from tests.random_instances import random_instances
from tests.real_instances import real_instances


def generate(data_source, rules):
    elections = {'election_id': [], 'voters': [], 'projects': [], 'budget': []}
    times = {'election_id': [], 'rule_id': [], 'time': []}
    if data_source == 'random':
        source = random_instances(10000, 1000, 300000, 3, 70)
    elif data_source == 'real':
        source = real_instances()
    else:
        raise ValueError("Unknown data source!")
    try:
        stop = False
        print("Starting testing")
        for n, E in enumerate(source):
            if n > 325:
                election_id = E.election_id
                print(f"Testing election {n}")
                elections['election_id'].append(election_id)
                elections['voters'].append(len(E.voters))
                elections['projects'].append(len(E.projects))
                elections['budget'].append(E.budget)
                for (rule, rule_id) in rules:
                    estimate = None
                    if rule_id == 0:
                        estimate = 4.911e-7 * len(E.projects) * E.budget
                    elif rule_id == 1:
                        estimate = 4.322e-8 * len(E.projects) * E.budget
                    elif rule_id == 2:
                        estimate = 2.327e-8 * len(E.voters) * len(E.projects) ** 2
                    if estimate <= 600:
                        start = timeit.default_timer()
                        ik_result = rule(E)
                        end = timeit.default_timer()
                        time = end - start
                        print(f"Finished on rule {rule_id}")
                        times['election_id'].append(election_id)
                        times['rule_id'].append(rule_id)
                        times['time'].append(time)
        raise KeyboardInterrupt
    except:
        try:
            path = f"benchmarks_{data_source}.xlsx"
            with pd.ExcelWriter(path, mode="a", if_sheet_exists='overlay', engine="openpyxl") as writer:

                election_frame = pd.DataFrame(elections)
                times_frame = pd.DataFrame(times)
                book = load_workbook(path)
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                election_frame.to_excel(writer, startrow=writer.sheets['Elections'].max_row, header=False, index=False,
                                        sheet_name="Elections")
                times_frame.to_excel(writer, startrow=writer.sheets['Times'].max_row, header=False, index=False,
                                     sheet_name="Times")
        except PermissionError:
            t = timeit.default_timer()
            path = f"benchmarks_{data_source}_{t}.xlsx"
            with pd.ExcelWriter(path, if_sheet_exists='overlay', engine="openpyxl") as writer:

                election_frame = pd.DataFrame(elections)
                times_frame = pd.DataFrame(times)
                book = load_workbook(path)
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                election_frame.to_excel(writer, startrow=writer.sheets['Elections'].max_row, header=False, index=False,
                                        sheet_name="Elections")
                times_frame.to_excel(writer, startrow=writer.sheets['Times'].max_row, header=False, index=False,
                                     sheet_name="Times")


if __name__ == "__main__":
    generate('real', [
        (interval_knapsack_projects_reversed, 2),
        (interval_knapsack_projects, 1)  # ,
        # (naive_interval_knapsack_projects, 0)
    ])
